import streamlit as st
import subprocess

@st.cache_resource
def install_playwright():
    subprocess.run(["playwright", "install", "chromium"])

install_playwright()

import os
import asyncio
from playwright.async_api import async_playwright
import datetime
import base64

from io import BytesIO

# Librerías para generación de Excel Técnico
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="LOTO OS Engine Matrix v4", layout="wide", initial_sidebar_state="collapsed")

# --- ESTILOS CSS ---
st.markdown("""
    <style>
    .custom-section-header {
        background-color: #B51E2D;
        color: white;
        padding: 10px;
        border-radius: 5px;
        font-size: 18px;
        font-weight: bold;
        margin-bottom: 15px;
        margin-top: 20px;
    }
    .mode-banner {
        padding: 15px;
        border-radius: 8px;
        color: white;
        font-weight: bold;
        text-align: center;
        margin: 10px 0;
        font-size: 16px;
    }
    .card-danger {
        background-color: #fee2e2;
        border-left: 5px solid #ef4444;
        padding: 10px;
        margin: 10px 0;
        border-radius: 4px;
    }
    .final-banner-premium {
        background-color: #B51E2D;
        color: white;
        padding: 20px;
        border-radius: 10px;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .final-banner-premium .label {
        font-size: 14px;
        text-transform: uppercase;
        opacity: 0.8;
    }
    .final-banner-premium .value {
        font-size: 32px;
        font-weight: 900;
        margin-top: 5px;
    }
    </style>
""", unsafe_allow_html=True)

# --- FUNCIONES AUXILIARES ---
def get_base64_image(file):
    if file is None: return None
    try:
        return base64.b64encode(file.getvalue()).decode()
    except:
        return None

def file_to_base64(path):
    if not os.path.exists(path): return None
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return None

import streamlit as st
from PIL import Image
import io

# --- 1. CONFIGURACIÓN E INICIALIZACIÓN ---
# Asegúrate de tener arca.png en la carpeta del proyecto
logo_path = "arca.png" 

st.title("IDENTIFICACIÓN DE MODO DE INTERVENCIÓN Y ANÁLISIS DE RIESGOS")

# --- 2. DATOS DESPLEGABLES ---
datos_maquinas = ["Aéreos", "Agrupadora", "Alimentador de cartón", "Alimentador de tapa", "Almacén", "Almacén de paleta", "Analizador", "Aplicador", "Aplicador de película", "Bailer", "Banda", "Batch Blend", "Bifurcador", "Bomba de incendios", "Bombas", "Bombeo", "Brazo", "Buffer", "Caldera", "Calentador", "Canulador", "Capsuladora", "Caracol", "Carboblending", "Carbocooler", "Carbonatador", "Carborproporcionador", "Chiller", "Codificador", "Codificador de lata", "Colocador de cánula", "Colocadora", "Combinador de paquetes", "Compactador", "Compresor", "Condensador", "Controlador de etiquetas", "Controlador de línea", "Coronado", "Cuarto frío", "Decapsulador", "Desempacadora", "Desencajonadora", "Desenfardadora", "Desenroscadora", "Desetiquetadora", "Despaletizadora", "Destapadora", "Disolución continua", "Disolución de carbón activado", "Distribuidor", "Divergente", "Divisor", "Divisora de Paquetes", "Dosificador", "Elevador", "Elevador de Tapas", "Empacadora", "Emplayadora", "Encajonadora", "Encogedora", "Enjuagadora", "Ensobretadora", "Entrada", "Envasadora", "Envolvedora", "Equipo Auxiliar", "Equipo TA", "Esmeril", "Estación", "Etiquetadora", "Extractor", "Filtro", "Flejadora", "Formadora", "Helix", "Hicone Applicator", "Homogenizador", "Horizontal Bailer", "Horno", "Humas", "Impresora", "Inspector", "Inspector de Botellas", "Inspector de Cajas", "Inspector de Etiquetas", "Inspector de Máquina de Lata", "Inspector de Nivel", "Inspector de Vacío", "Inspector Electrónico", "Intercambiador", "Intercambiador de Calor", "Inyector", "Lámpara", "Lantech 1", "Lavadora", "Lavadora de cajas", "Lavadora de garrafón", "Lente", "LGV", "Llenadora", "Llenadora Tetrapack", "Llenadora y Rinser", "Magazine de Cajas", "Magazine de Pallets", "Marmita", "Mesa", "Mesa Acumuladora", "Mesa de Descarga", "Mezclador", "Mixer", "Motor", "Nitrógeno", "Ordenador", "Ordenador de cánulas", "Ósmosis", "Ozono", "Paletizadora", "Pasteurizador", "Polipasto", "Posicionador de Paquetes", "PP D001", "PP D002", "PP D003", "PP D004", "PP D005", "PP D006", "PP D007", "PP D008", "PP D009", "PP D010", "PP D011", "PP D012", "Precintador", "Prensa", "Preparador", "Proporcionador", "Rampa", "Rechazo", "Refrigeración", "Revisador", "Rinser", "Robot", "Rolador", "Roscadora", "Ruta", "Salida", "Secador", "Separador", "Sierra", "Sistema de NH3", "Sistema Ergonómico", "Sistema PTAN", "Soldadora", "Soplador", "Straw", "Tablero Eléctrico", "Taladro", "Tanque", "Tanque de disolución", "Tanque de dosage", "Taponadora", "Termoencogible", "Termoformadora", "Tolva", "Torre Enfriamiento", "Transformador", "Transporte de Pallets", "Transporte de Packs", "Transporte de Cajas", "Transportador", "Transporte de Botellas", "Transporte de Preformas", "Trechas", "Triturador", "Túnel de Enfriamiento", "Vertical Bailer", "Pulmón"]

opciones_tareas = {
    "Operaciones normales (0 acceso): Protecciones instaladas de conformidad con OSH-RQ-185": [
        "Operación normal de equipos", "Monitoreo de parámetros operacionales", "Supervisión de producción", 
        "Verificación de alarmas y tendencias", "Cambio de setpoints autorizados", "Arranque normal de equipos", 
        "Detención normal de equipos", "Operación desde HMI"
    ],
    "Intervención Menor (trabajo a través o dentro de áreas protegida): rutinaria, repetitiva e integral": [
        "Acomodo de cartón dañado", "Acomodo de paquetes fuera de posición", "Acomodo de tarima fuera de posición", "Ajuste de clutch", "Ajuste de espreas", "Ajuste de guías", "Ajuste de gusano de entrada", "Ajuste de pata de gallo", "Ajuste de placas", "Ajuste de presión de aire", "Ajuste de sensores", "Ajuste de tiempos de manejo", "Ajuste de torque", "Ajuste de válvulas", "Ajuste de ventosas", "Ajuste por cambio de presentación", "Alineación de sensores", "Aplicación de vortex", "Cambio de bandas", "Cambio de baleros", "Cambio de botella dañada", "Cambio de componentes mecánicos", "Cambio de copas", "Cambio de filtros", "Cambio de guías", "Cambio de manejos", "Cambio de rollo de termoencogible", "Cambio de sistema de arrastre", "Cambio de válvulas", "Cambio de tubos de ventila", "Colocación de tulipas o botellas falsas", "Corte de producto", "Corte manual de línea", "Destrabe de botellas", "Destrabe de cajas", "Inspección ultrasónica", "Levantamiento de botellas caídas", "Levantamiento de paquetes atorados", "Limpieza de áreas de transporte", "Limpieza de cámaras interiores", "Limpieza de componentes mecánicos", "Limpieza de cortinas", "Limpieza de filtros y conexiones", "Limpieza de mesa de corte", "Limpieza de mesa de transferencia", "Limpieza de sensores", "Limpieza de sopladores", "Limpieza de ventiladores", "Limpieza interior de equipos", "Lubricación centralizada", "Lubricación de bandas", "Lubricación de elevadores", "Lubricación de rodillos", "Lubricación de sistemas mecánicos", "Lubricación general", "Operación manual de válvulas", "Predictivo por ultrasonido", "Procedimiento por explosión de botella", "Reemplazo de componentes desgastados", "Reposición de componentes de transportador", "Retiro de acumulación de cartón", "Retiro de acumulación de tapas", "Retiro de botellas defectuosas", "Retiro de botellas sin envolver", "Retiro de etiquetas acumuladas", "Retiro de etiquetas mal aplicadas", "Retiro de plástico y residuos", "Retiro de producto defectuoso", "Retiro de tarima dañada", "Revisión de bandas", "Revisión de cangilones", "Revisión de componentes mecánicos", "Revisión de conexiones", "Revisión de encoder", "Revisión de estrellas de sujeción", "Revisión de filtros", "Revisión de guardas de seguridad", "Revisión de hornilla", "Revisión de intercambiadores", "Revisión de reductores", "Revisión de rodillos", "Revisión de sensores", "Revisión de separadores", "Revisión de sistemas de arrastre", "Revisión de sistemas de seguridad", "Revisión de sopladores", "Revisión de tensores", "Revisión de transportadores", "Revisión de túnel de encogimiento", "Revisión de unidades de mantenimiento", "Revisión de ventiladores", "Verificación de bandas deslizantes", "Verificación de componentes eléctricos", "Verificación de conexiones", "Verificación de guardas protectoras", "Verificación de inocuidad", "Verificación de micro switches", "Verificación de presión", "Verificación de seguridad operacional", "Verificación de sistemas neumáticos", "Verificación de transportadores"
    ]
}

# --- 3. INTERFAZ ---
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="custom-section-header">Información de Sitio</div>', unsafe_allow_html=True)
    lista_sitios = [
        "Planta Tucumán", "Planta Formosa", "Planta Salta",
        "CEDI Corrientes", "CEDI Goya", "CEDI Posadas", "CEDI Resistencia", "CEDI Saenz Peña",
        "Ingenio Famaillá", "Ingenio Bella Vista", "CEDI J.V.Gonzalez", "CEDI Ledesma",
        "CEDI Metán", "CEDI Orán", "CEDI Jujuy", "CEDI San Pedro", "CEDI Sgo Del Estero",
        "CEDI Tartagal", "CEDI Chilecito", "CEDI Concepción", "CEDI La Rioja", "CEDI Catamarca"
    ]
    sitio = st.selectbox("Sitio", lista_sitios)
    
    # Clasificación de Negocio
    negocio = "Bebidas"
    tipo_sitio = "Planta" if "Planta" in sitio else "CEDI"
    if "Ingenio" in sitio:
        negocio = "Ingenios"
        tipo_sitio = "Ingenio"
        
    area_sector = st.selectbox("Área / Sector", [
        "Producción", "Mantenimiento", "Calidad", "Servicios Auxiliares", 
        "Expedición", "Comercial", "Logística", "Automotor"
    ])
    
    linea = "N/A"
    maquina = "N/A"
    subsector_equipo = "N/A"
    
    if area_sector == "Producción":
        # Lógica de líneas según planta
        if sitio == "Planta Salta":
            linea = st.selectbox("Línea de Embotellado", ["Línea 01", "Línea 02", "Línea 03", "Línea 04", "Línea 05", "Línea 06", "Línea Dual Pack", "Línea 1000 Tetra", "Línea 200 Tetra", "Línea Bidones", "Sorting", "BBOX"])
        elif sitio == "Planta Tucumán":
            linea = st.selectbox("Línea de Embotellado", ["Línea 01", "Línea 02", "Línea 03", "Línea 04", "Línea 05", "Línea 06", "Sorting", "BBOX"])
        else:
            linea = st.selectbox("Línea de Embotellado", ["Línea 01", "Línea 02", "Línea 03", "Línea 04"])
        maquina = st.selectbox("Máquina", datos_maquinas)
    else:
        subsector_equipo = st.text_input("Subsector / Equipo o Sistema")
    
    st.markdown('<div class="custom-section-header">Datos de Equipo</div>', unsafe_allow_html=True)
    fabricante = st.text_input("Fabricante")
    modelo = st.text_input("Modelo / Marca")
    anio = st.number_input("Año de fabricación", min_value=1950, max_value=2030, step=1)

with col1:
    st.markdown('<div class="custom-section-header">Clasificación de Tarea</div>', unsafe_allow_html=True)
    clasificacion = st.radio("Seleccione el modo:", list(opciones_tareas.keys()) + ["Mantenimiento", "Equipo energizado (modo jog o enseñanza del equipo)"])

    tareas_predefinidas = []
    if clasificacion in opciones_tareas:
        tareas_predefinidas = st.multiselect("Seleccione tarea(s):", opciones_tareas[clasificacion])
    
    tareas_manuales = st.text_area("Identifique tareas adicionales (Enter para nuevas líneas):")
    
    # Combinar tareas para el procesamiento posterior
    lista_tareas_combinadas = []
    if tareas_predefinidas:
        lista_tareas_combinadas.extend(tareas_predefinidas)
    if tareas_manuales:
        lista_tareas_combinadas.extend([t.strip() for t in tareas_manuales.split("\n") if t.strip()])
    
    frecuencia = st.selectbox("Frecuencia promedio", [
        "De 15 a 20 veces por hora", "De 5 a 20 veces por turno", "De 1 a 15 veces por turno",
        "De 3 a 10 veces por turno", "De 3 a 6 veces por turno", "De 2 a 4 veces por turno",
        "De 1 a 5 veces por turno", "De 1 a 3 veces por turno", "De 1 a 2 veces por turno",
        "De 1 a 2 veces por día", "De 3 a 4 veces por semana", "De 1 a 3 veces por semana",
        "1 vez por semana", "De 1 a 3 veces por mes", "De 1 a 2 veces por mes",
        "1 vez cada 2 meses", "1 vez cada 3 meses", "1 vez cada 4 meses",
        "1 vez cada 6 meses", "1 vez al año", "1 vez cada 2 años",
        "1 vez cada 3 años", "1 vez cada 4 años", "1 vez cada 5 años"
    ])
    duracion = st.selectbox("Duración promedio", [
        "1 minuto", "2 minutos", "3 minutos", "5 minutos", "10 minutos", "15 minutos",
        "25 minutos", "30 minutos", "1 hora", "2 horas", "3 horas", "4 horas",
        "5 horas", "7 horas", "8 horas", "De 9 a 12 horas", "Más de 12 horas"
    ])
    
    tareas = lista_tareas_combinadas

with col2:
    st.markdown('<div class="custom-section-header">Carga de Foto</div>', unsafe_allow_html=True)
    equip_file = st.file_uploader("Subir Foto del Equipo", type=["png", "jpg", "jpeg"])
    if equip_file:
        st.image(equip_file, caption="Equipo Seleccionado", use_container_width=True)

    st.markdown('<div class="custom-section-header">Identificación de Modo Inicial de Intervención</div>', unsafe_allow_html=True)
    nivel_acceso = st.radio(
        "Nivel de Acceso / Exposición Requerido",
        [
            "La tarea requiere que el equipo esté energizado mientras se trabaja en o alrededor de zonas peligrosas de equipo",
            "La tarea incluye desmontaje o contacto con partes energizadas",
            "La tarea requiere ACCESO CORPORAL COMPLETO (trabajar DENTRO DE GUARDA INTERBLOQUEADO)",
            "La tarea requiere ACCESO CORPORAL PARCIAL (trabajar A TRAVÉS DE GUARDA INTERBLOQUEADO)",
            "La tarea no requiere NINGÚN TIPO DE ACCESO CORPORAL"
        ]
    )
    
    # --- LÓGICA MODO INICIAL ---
    modo_inicial = "Modo 0"
    if "equipo esté energizado" in nivel_acceso:
        modo_inicial = "Modo 4"
    elif "desmontaje o contacto" in nivel_acceso:
        modo_inicial = "Modo 3"
    elif "ACCESO CORPORAL COMPLETO" in nivel_acceso:
        modo_inicial = "Modo 2"
    elif "ACCESO CORPORAL PARCIAL" in nivel_acceso:
        modo_inicial = "Modo 1"
        
    colores_modos = {
        "Modo 0": "#69C97F",
        "Modo 1": "#00B0F0",
        "Modo 2": "#00B0F0",
        "Modo 3": "#FF560E",
        "Modo 4": "#F30009"
    }
    color_ini = colores_modos.get(modo_inicial, "#B51E2D")
    st.markdown(f'<div class="mode-banner" style="background-color: {color_ini}; color: white;">Modo Inicial Identificado: {modo_inicial}</div>', unsafe_allow_html=True)
    
    # Mensajes de seguridad dinámicos
    mensajes_seguridad = {
        "Modo 0": "Asegúrese de que el ACCESO CERO, la protección de la máquina debe estar en su lugar y en buenas condiciones.",
        "Modo 1": "DETENGA LA MÁQUINA y trabaje a través de la GUARDA CON INTERLOCK.",
        "Modo 2": "DETENGA LA MÁQUINA y aplique 'BLOQUEO' para EVITAR el REINICIO.",
        "Modo 3": "DETENGA LA MÁQUINA y aplique LOTO COMPLETO para realizar la tarea.",
        "Modo 4": "MÁXIMA PRECAUCIÓN: Equipo energizado con acceso a zonas peligrosas."
    }
    
    if modo_inicial in ["Modo 1", "Modo 2"]:
        msg_inicial = "*Modo sujeto a validación en 5. VALIDACIÓN DEL MODO INICIAL"
    else:
        msg_inicial = mensajes_seguridad.get(modo_inicial, "")
        
    if msg_inicial:
        st.markdown(f'<div style="color: {color_ini}; font-weight: bold; border: 2px solid {color_ini}; padding: 10px; border-radius: 5px;">Acción Requerida: {msg_inicial}</div>', unsafe_allow_html=True)
        
    st.markdown('<div style="margin-top: 25px;"></div>', unsafe_allow_html=True)
    evaluador = st.text_input("Nombre y Apellido de Evaluador")
    puesto_evaluador = st.text_input("Puesto")
    fecha_evaluacion = st.date_input("Fecha de Evaluación", value=datetime.date.today())

# --- FUENTES DE ENERGÍA ---
st.markdown('<div class="custom-section-header">3. Fuentes de Energía(s) para la(s) Tarea(s) / Modo Seleccionado</div>', unsafe_allow_html=True)
fuentes_energia = [
    ("Eléctrica", "380V"), ("Mecánica", "N/A"), ("Hidráulica", "N/A"),
    ("Neumática", "6 bar"), ("Térmica", "N/A"), ("Química", "N/A"),
    ("Potencial", "N/A"), ("Magnética", "N/A"), ("Radiante", "N/A"),
    ("Presión de Agua", "N/A")
]

energias_seleccionadas = {}
cols_eng = st.columns(3)
for i, (nombre, val_def) in enumerate(fuentes_energia):
    with cols_eng[i % 3]:
        activado = st.checkbox(nombre, value=(i==0 or i==3), key=f"chk_{i}")
        if activado:
            if nombre == "Química":
                sustancias_base = ["Amoníaco (NH3)", "Gas carbónico (CO2)", "Soda cáustica (NaOH)", "Nitrógeno", "Gas Oil", "Gas Natural", "GLP"]
                seleccion_sustancias = st.multiselect("Identifique la sustancia(s) química(s):", options=sustancias_base + ["Otra (Manual)"], key="sel_quimica")
                
                detalles_quimica = []
                for s in seleccion_sustancias:
                    if s == "Otra (Manual)":
                        otra_s = st.text_input("Especifique otra sustancia:", key="otra_quimica")
                        mag_s = st.text_input(f"Magnitud para {otra_s}:", key="mag_otra_quimica")
                        if otra_s: detalles_quimica.append(f"{otra_s} ({mag_s})")
                    else:
                        mag_s = st.text_input(f"Magnitud para {s}:", key=f"mag_{s}")
                        detalles_quimica.append(f"{s} ({mag_s})")
                
                if detalles_quimica:
                    energias_seleccionadas[nombre] = " | ".join(detalles_quimica)
            else:
                magnitud = st.text_input(f"Magnitud {nombre}", value=val_def, key=f"mag_{i}")
                energias_seleccionadas[nombre] = magnitud

controles_por_categoria = {
    "Ingeniería": [
        "Interlocks / enclavamientos de puertas", "Sensores ópticos / cortinas de luz / barreras infrarrojas",
        "Paradas de emergencia", "Vallas fijas 360° / cerramientos / mallas metálicas",
        "Protecciones físicas de transmisiones y engranajes", "Plataformas, pasarelas y escaleras antideslizantes",
        "Barandas, rodapiés y travesaños intermedios", "Sistemas de presión positiva",
        "Ventilación industrial / extractores / ventiladores", "Sistemas automáticos de parada",
        "Sistemas automatizados de sanitación o inyección química", "Protección contra proyección de vidrios",
        "Puestas a tierra / diferenciales eléctricos / tensión segura 24V", "Iluminación LED y mejoras de iluminación",
        "Red fija contra incendios", "Alarmas sonoras y balizas luminosas", "Detectores de oxígeno / gases",
        "Separación física peatón–autoelevador", "Diseño ergonómico de puestos",
        "Carros ergonómicos y ayudas mecánicas", "Pisos y superficies antideslizantes",
        "Bandejas de contención y drenajes", "Limitadores de velocidad en autoelevadores"
    ],
    "Administrativas / Procedimientos": [
        "Procedimiento de Bloqueo Operacional", "Procedimiento LOTO / Control de Energías Peligrosas",
        "Procedimiento Protección de Máquinas", "Procedimientos seguros de limpieza, destrabe y sanitación",
        "Procedimientos para trabajos en altura", "Procedimientos eléctricos seguros",
        "Procedimientos para manejo de sustancias peligrosas", "Procedimientos de manejo de contratistas",
        "Permisos de trabajo", "Reglas que Salvan Vidas (RSV)", "Call To Action / SIF / HOP",
        "Programa de Seguridad Basado en el Comportamiento (PSBC)", "Charlas de 5 minutos",
        "Tarjetas de observación", "Reportes e investigación de incidentes",
        "Controles operacionales y rondas de seguridad", "Señalización de advertencia, prohibición y obligación",
        "Demarcación de sendas peatonales", "Reglas para peatones y autoelevadores",
        "Campañas preventivas (golpe de calor, alcoholismo, etc.)", "Simulacros y brigadas de emergencia",
        "Capacitación técnica y de seguridad", "Certificación de operadores y equipos",
        "Planes de mantenimiento preventivo y correctivo", "Estudios ambientales y mediciones"
    ],
    "EPP": [
        "Calzado de seguridad", "Anteojos de seguridad", "Protector facial", "Guantes anticorte",
        "Guantes dieléctricos", "Guantes de PVC", "Mamelucos impermeables", "Protección respiratoria",
        "Protectores auditivos", "Arnés anticaídas", "Chalecos reflectivos", "Casco de seguridad", "Mangas anticorte"
    ],
    "Salud Ocupacional": [
        "Estudios médicos preventivos y periódicos", "Exámenes preocupacionales", "Estudios psicológicos y neurológicos",
        "Vigilancia de salud ocupacional", "Pausas activas / descansos reglamentados", "Control de carga térmica",
        "Control audiométrico", "Seguimiento médico de exposición", "Programas de bienestar y salud mental",
        "Prevención de alcoholismo y drogas"
    ],
    "Gestión de Emergencias": [
        "Planes de emergencia", "Simulacros de evacuación", "Formación de brigadas", "Respuesta a emergencias químicas",
        "Procedimientos ante incendios", "Sistemas de evacuación y señalización de salidas"
    ],
    "Seguridad Patrimonial y Tránsito": [
        "Cámaras de monitoreo", "Guardia privada", "Control de circulación vehicular",
        "Alarmas de retroceso en autoelevadores", "Balizas reflectivas", "Señalización vial interna",
        "Sendas peatonales segregadas", "Control de velocidad vehicular"
    ],
    "Gestión Ergonómica": [
        "Rotación de tareas", "Banquetas ergonómicas", "Regulación de altura de puestos",
        "Ayudas mecánicas para manipulación", "Estudios ergonómicos", "Rediseño de puestos de trabajo",
        "Disminución de manipulación manual de cargas"
    ]
}

peligros_por_categoria = {
    "Mecánicos": [
        "Atrapamiento en partes móviles", "Enredos con elementos rotativos", "Aplastamiento", "Cizallamiento",
        "Golpes contra objetos fijos o móviles", "Proyección de partículas o fragmentos", "Caída de objetos",
        "Contacto con superficies cortantes o punzantes", "Energía mecánica almacenada", "Movimiento inesperado de equipos",
        "Fallas o liberación de presión mecánica"
    ],
    "Eléctricos": [
        "Contacto eléctrico directo", "Contacto eléctrico indirecto", "Arco eléctrico", "Energía eléctrica almacenada",
        "Baja tensión (220V / 380V)", "Equipos energizados"
    ],
    "Físicos": [
        "Ruido elevado (>85 dBA)", "Temperaturas extremas (calor / frío)", "Iluminación insuficiente",
        "Vibraciones", "Radiación no ionizante", "Superficies calientes o frías"
    ],
    "Otros": [
        "Resbalones y tropiezos", "Caídas al mismo nivel", "Caídas a distinto nivel", "Superficies irregulares",
        "Orden y limpieza deficientes", "Deficiencias en accesos o circulación", "Incendio", "Explosión"
    ],
    "Ergonómicos": [
        "Manipulación manual de cargas", "Sobreesfuerzo físico", "Posturas forzadas", "Movimientos repetitivos",
        "Trabajos prolongados de pie", "Diseño inadecuado del puesto de trabajo"
    ],
    "Químicos": [
        "Inhalación de sustancias químicas", "Contacto químico con piel u ojos", "Exposición a vapores, gases o nieblas",
        "Sustancias corrosivas", "Sustancias inflamables", "Derrames químicos"
    ],
    "Transporte / Tránsito": [
        "Choques o colisiones vehiculares", "Interacción peatón–autoelevador", "Estado deficiente de caminos o rutas",
        "Atropellamiento", "Maniobras inseguras"
    ],
    "Psicosociales": [
        "Carga excesiva de trabajo", "Fatiga laboral", "Estrés laboral", "Violencia o acoso laboral",
        "Robos o asaltos", "Consumo de alcohol o sustancias", "Condiciones personales de salud que afecten la tarea"
    ],
    "Especiales": [
        "Espacios confinados", "Trabajo en altura", "Trabajos en caliente", "Atmósferas peligrosas",
        "Deficiencia de oxígeno", "Riesgo biológico", "Condiciones climáticas severas"
    ]
}

dict_prob = {
    0.1: "0.1 - Casi imposible: secuencia o consecuencia prácticamente imposible",
    0.5: "0.5 - Coincidencia extremadamente remota pero concebible",
    1.0: "1.0 - Sería una coincidencia remotamente posible. Se sabe que ha ocurrido. (1%)",
    3.0: "3.0 - Sería una secuencia o coincidencia rara: 10% (Último año)",
    6.0: "6.0 - Es posible: nada extraño tiene una probabilidad del 50% (Últimos 6 meses)",
    10.0: "10.0 - Convicción: es el resultado más probable y esperado"
}
dict_exp = {
    0.5: "0.5 - Muy raramente el riesgo existe, sin antecedente de personal expuesto",
    1.0: "1.0 - Raramente: el riesgo existe y se tiene antecedentes de personal expuesto",
    2.0: "2.0 - Inusual: hasta una vez al año",
    3.0: "3.0 - Ocasionalmente: una vez a la semana o una vez al mes",
    6.0: "6.0 - Frecuentemente: Aproximadamente una o dos veces al día",
    10.0: "10.0 - Continuamente: La situación ocurre continuamente o muchas veces al día"
}
dict_cons = {
    1.0: "1.0 - Heridas leves, contusiones, golpes",
    5.0: "5.0 - Lesiones incapacitantes",
    15.0: "15.0 - Lesiones muy graves, amputaciones, invalidez permanente",
    25.0: "25.0 - Una muerte",
    50.0: "50.0 - Varias muertes o daños materiales graves",
    100.0: "100.0 - Catastróficas: numerosas muertes, grandes daños"
}

def clasificar_fine(gp):
    if gp > 401: return "Riesgo Inminente", "#EF4444"
    elif 201 <= gp <= 400: return "Riesgo Alto", "#F97316"
    elif 71 <= gp <= 200: return "Riesgo Notable", "#EAB308"
    elif 21 <= gp <= 70: return "Riesgo Moderado", "#3B82F6"
    else: return "Riesgo Aceptable", "#10B981"

# --- 4. EVALUACIÓN COMPLETA DE RIESGOS ---
st.markdown('<div class="custom-section-header">4. Identificación de Peligros y Evaluación de Riesgos (W. Fine)</div>', unsafe_allow_html=True)

categorias_seleccionadas = st.multiselect("Seleccione Categoría(s) de Peligro:", options=list(peligros_por_categoria.keys()))

lista_peligros_final = []
for cat in categorias_seleccionadas:
    pels = st.multiselect(f"Peligros en categoría {cat}:", options=peligros_por_categoria[cat], key=f"sel_pel_{cat}")
    lista_peligros_final.extend(pels)
    p_manual = st.text_input(f"Peligros adicionales manuales para {cat}:", key=f"p_man_{cat}")
    if p_manual:
        lista_peligros_final.append(f"{cat}: {p_manual}")

evaluaciones_fine = []

if lista_peligros_final:
    for i, pel in enumerate(lista_peligros_final):
        st.markdown(f'<div class="card-danger"><strong>⚠️ Peligro #{i+1}:</strong> {pel}</div>', unsafe_allow_html=True)
        
        c_i1, c_i2, c_i3 = st.columns(3)
        with c_i1: p_i = st.selectbox(f"Probabilidad (Inherente) #{i+1}", options=list(dict_prob.keys()), format_func=lambda x: dict_prob[x], key=f"p_i_{i}", index=3)
        with c_i2: e_i = st.selectbox(f"Exposición (Inherente) #{i+1}", options=list(dict_exp.keys()), format_func=lambda x: dict_exp[x], key=f"e_i_{i}", index=4)
        with c_i3: c_i = st.selectbox(f"Consecuencia (Inherente) #{i+1}", options=list(dict_cons.keys()), format_func=lambda x: dict_cons[x], key=f"c_i_{i}", index=2)
        
        gp_inherente = round(p_i * e_i * c_i, 1)
        clase_i, color_i = clasificar_fine(gp_inherente)
        st.markdown(f'<p style="color:{color_i}; font-size:12px; margin-top:-10px;"><strong>GP Inherente: {gp_inherente} ({clase_i}) [P:{p_i} * E:{e_i} * C:{c_i}]</strong></p>', unsafe_allow_html=True)
        
        st.markdown(f"**Asignar Medidas de Control para Peligro #{i+1}:**")
        cat_ctrl_sel = st.multiselect(f"Seleccione Categoría(s) de Control #{i+1}:", options=list(controles_por_categoria.keys()), key=f"cat_ctrl_{i}")
        
        lista_ctrl_final = []
        for c_cat in cat_ctrl_sel:
            c_sel = st.multiselect(f"Medidas en {c_cat} #{i+1}:", options=controles_por_categoria[c_cat], key=f"sel_ctrl_{c_cat}_{i}")
            lista_ctrl_final.extend(c_sel)
            cs_manual = st.text_input(f"Medidas adicionales manuales para {c_cat} (Peligro #{i+1}):", key=f"cs_man_{c_cat}_{i}")
            if cs_manual:
                lista_ctrl_final.append(cs_manual)
        
        controles_seleccionados = lista_ctrl_final
        
        c_r1, c_r2, c_r3 = st.columns(3)
        with c_r1: p_r = st.selectbox(f"Probabilidad (Residual) #{i+1}", options=list(dict_prob.keys()), format_func=lambda x: dict_prob[x], key=f"p_r_{i}", index=1)
        with c_r2: e_r = st.selectbox(f"Exposición (Residual) #{i+1}", options=list(dict_exp.keys()), format_func=lambda x: dict_exp[x], key=f"e_r_{i}", index=2)
        with c_r3: c_r = st.selectbox(f"Consecuencia (Residual) #{i+1}", options=list(dict_cons.keys()), format_func=lambda x: dict_cons[x], key=f"c_r_{i}", index=1)
        
        gp_residual = round(p_r * e_r * c_r, 1)
        clase_r, color_r = clasificar_fine(gp_residual)
        st.markdown(f'<p style="color:{color_r}; font-size:12px; margin-top:-10px;"><strong>GP Residual Mitigado: {gp_residual} ({clase_r}) [P:{p_r} * E:{e_r} * C:{c_r}]</strong></p>', unsafe_allow_html=True)
        st.markdown('---')
        
        evaluaciones_fine.append({
            "peligro": pel, "controles": ", ".join(controles_seleccionados),
            "p_i": p_i, "e_i": e_i, "c_i": c_i, "gp_i": gp_inherente, "clase_i": clase_i,
            "p_r": p_r, "e_r": e_r, "c_r": c_r, "gp_r": gp_residual, "clase_r": clase_r, "color_r": color_r
        })

# --- 5. VALIDACIÓN DEL MODO SEGURO (Solo para Modo 1 y 2) ---
modo_final = modo_inicial
riesgo_no_aceptable = any(item["gp_r"] > 20 for item in evaluaciones_fine)

if modo_inicial in ["Modo 1", "Modo 2"]:
    st.markdown('<div class="custom-section-header">5. Validación del Modo Inicial</div>', unsafe_allow_html=True)
    st.warning(f"Validación requerida para {modo_inicial}")
    
    if modo_inicial == "Modo 1":
        pregunta = "¿Se garantiza control fiable para prevenir rearranque? Enclavamiento (ISO PLd/PLe, ANSI Cat. 3/4)"
        control_fiable = st.radio(pregunta, ["Sí", "No"], index=0)
        if riesgo_no_aceptable or control_fiable == "No":
            modo_final = "Modo 3"
            st.error("⚠️ Validación Fallida: Se requiere MODO FINAL 3 debido a riesgo residual > 20 o falta de control fiable.")
            
    elif modo_inicial == "Modo 2":
        pregunta = "¿Se garantiza control fiable para prevenir rearranque? Enclavamiento (ISO PLd/PLe, ANSI Cat. 3/4) + bloqueo físico de puerta o sistema de Llave Atrapada (Trapped o Fortress Key)"
        control_fiable = st.radio(pregunta, ["Sí", "No"], index=0)
        if riesgo_no_aceptable or control_fiable == "No":
            modo_final = "Modo 3"
            st.error("⚠️ Validación Fallida: Se requiere MODO FINAL 3 debido a riesgo residual > 20 o falta de control fiable.")

# --- 6. CONCLUSIÓN ---
num_concl = "5" if modo_inicial in ["Modo 0", "Modo 3", "Modo 4"] else "6"
st.markdown(f'<div class="custom-section-header">{num_concl}. Conclusiones / Observaciones</div>', unsafe_allow_html=True)
col_aud1, col_aud2 = st.columns(2)

with col_aud1:
    medidas_conclusiones = st.text_area("Observaciones", value="")

with col_aud2:
    color_fin = colores_modos.get(modo_final, "#B51E2D")
    st.markdown(f"""
        <div class="mode-banner" style="background-color: {color_fin}; font-size: 20px; padding: 20px; color: white;">
            <div style="font-size: 12px; text-transform: uppercase; opacity: 0.8;">Modo Final de Intervención</div>
            {modo_final.upper()}</div>
    """, unsafe_allow_html=True)
    
    msg_final = mensajes_seguridad.get(modo_final, "")
    if msg_final:
        st.markdown(f'<div style="color: {color_fin}; font-weight: bold; border: 2px solid {color_fin}; padding: 15px; border-radius: 5px; text-align: center;">Instrucción Final: {msg_final}</div>', unsafe_allow_html=True)

# --- PLANTILLA HTML AVANZADA ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page { size: A4 portrait; margin: 10mm 8mm; }
        body { font-family: 'Bahnschrift', Arial, sans-serif; color: #0f172a; margin: 0; line-height: 1.2; font-size: 8.5pt; }
        .header-table { width: 100%; border-collapse: collapse; margin-bottom: 12px; border: 1.5pt solid #000; }
        .header-table td { border: 0.5pt solid #000; vertical-align: middle; padding: 5px; }
        .header-logo { width: 20%; text-align: center; background-color: #fff; }
        .header-title { width: 55%; text-align: center; background-color: #000; color: #fff; }
        .header-title h1 { font-size: 12pt; margin: 2px 0; text-transform: uppercase; }
        .header-title h2 { font-size: 10pt; margin: 2px 0; font-weight: normal; color: #ccc; }
        .header-title h3 { font-size: 11pt; margin: 2px 0; color: #fff; background-color: #444; padding: 2px; }
        .header-info { width: 25%; font-size: 8pt; background-color: #eee; }
        .header-info table { width: 100%; border-collapse: collapse; }
        .header-info td { border: none; padding: 2px 5px; }
        
        .section { margin-bottom: 10px; border: 1px solid #B51E2D; border-radius: 4px; overflow: hidden; page-break-inside: avoid; }
        .section-header { background-color: #B51E2D; color: #ffffff; padding: 6px 10px; font-size: 9pt; font-weight: bold; text-transform: uppercase; }
        
        .grid-layout { width: 100%; border-collapse: collapse; }
        .grid-layout td { padding: 6px 10px; border: 1px solid #e2e8f0; vertical-align: top; }
        
        .matrix-table { width: 100%; border-collapse: collapse; font-size: 8pt; margin-top: 5px; }
        .matrix-table th { background: #444444; color: white; padding: 5px; text-align: center; font-weight: bold; border: 1px solid #444444; }
        .matrix-table td { padding: 6px; border: 1px solid #cbd5e1; text-align: center; }
        .left-align { text-align: left !important; }
        
        .img-container { text-align: center; padding: 10px; }
        .img-container img { max-width: 100%; height: auto; max-height: 160px; border: 1px solid #cbd5e1; border-radius: 4px; }
        
	        .banner-final-pdf { padding: 10px; text-align: center; border: 2px solid #B51E2D; border-radius: 4px; margin-top: 12px; page-break-inside: avoid; color: white; }
	    </style>
	</head>
	<body>
    <table class="header-table">
        <tr>
            <td class="header-logo">__LOGO_HTML__</td>
            <td class="header-title">
                <h1>CONTROL DE ENERGÍAS PELIGROSAS</h1>
                <h2>__NEGOCIO__ - __PLANTA__</h2>
                <h3>MODOS DE INTERVENCIÓN Y EVALUACIÓN DE RIESGOS</h3>
            </td>
            <td class="header-info">
                <table>
                    <tr><td><strong>CÓDIGO:</strong></td><td></td></tr>
                    <tr><td><strong>REVISIÓN:</strong></td><td></td></tr>
                    <tr><td><strong>FECHA:</strong></td><td>__FECHA_ACTUAL__</td></tr>
                </table>
            </td>
        </tr>
    </table>

    <div class="section">
        <div class="section-header">1. IDENTIFICACIÓN DE EQUIPO Y TAREAS</div>
        <table class="grid-layout">
            <tr>
                <td style="width: 50%;">
                    <strong>Negocio:</strong> __NEGOCIO__ | <strong>Sitio:</strong> __PLANTA__<br>
                    <strong>Ubicación:</strong> __AREA__ — __LINEA__<br>
                    <strong>Equipo/Máquina:</strong> __EQUIPO_DESC__<br>
                    <strong>Fabricante:</strong> __FABRICANTE__ | <strong>Modelo:</strong> __MODELO__ | <strong>Año:</strong> __ANIO__<br>
                    <strong>Clasificación de tarea:</strong> __CLASIFICACION__<br>
                    <strong>Frecuencia:</strong> __FRECUENCIA__ | <strong>Duración:</strong> __DURACION__<br>
                    <strong>Grupo de Tareas:</strong><br>__TAREAS__
                </td>
                <td style="width: 50%; text-align: center;" class="img-container">
                    <strong>REGISTRO VISUAL DEL EQUIPO:</strong><br>
                    __FOTO_EQUIPO__
                </td>
            </tr>
        </table>
    </div>

    <div class="section">
        <div class="section-header">2. IDENTIFICACIÓN DE MODO INICIAL DE INTERVENCIÓN</div>
        <table class="grid-layout">
            <tr>
                <td>
                    <strong>Nivel de Acceso / Exposición Requerido:</strong> __NIVEL_ACCESO__<br>
                    <strong>Modo Inicial Identificado:</strong> __MODO_INICIAL__<br>
                    <strong>Acción Requerida:</strong> __ACCION_REQUERIDA__<br>
                    <strong>Evaluador:</strong> __EVALUADOR_FULL__
                </td>
            </tr>
        </table>
    </div>

    <div class="section">
        <div class="section-header">3. FUENTES DE ENERGÍA(S) PARA LA(S) TAREA(S) / MODO SELECCIONADO</div>
        <table class="grid-layout">
            <tr>
                <td>
                    <strong>Fuentes de Energía Identificadas:</strong><br>
                    __ENERGIA__
                </td>
            </tr>
        </table>
    </div>

    <div class="section">
        <div class="section-header">4. Identificación de Peligros y Evaluación de Riesgos (W. Fine)</div>
        <table class="matrix-table">
            <thead>
                <tr>
                    <th rowspan="2">Peligro Identificado</th>
                    <th colspan="4">Evaluación de Riesgos (Inherente)</th>
                    <th rowspan="2">Medidas de control</th>
                    <th colspan="4">Evaluación de Riesgos (Residual)</th>
                </tr>
                <tr>
                    <th>P</th><th>E</th><th>C</th><th>GP</th>
                    <th>P</th><th>E</th><th>C</th><th>GP</th>
                </tr>
            </thead>
            <tbody>
                __FILAS_MATRIZ__
            </tbody>
        </table>
        <div style="font-size: 7pt; padding: 5px; background: #f8fafc; border-top: 1px solid #cbd5e1;">
            <strong>Leyenda:</strong> P: Probabilidad | E: Exposición | C: Consecuencia | GP: Grado de Peligrosidad (P x E x C)
        </div>
    </div>

    __SECCION_VALIDACION__

    <!-- Sección Árbol Eliminada -->

    <div class="section">
        <div class="section-header">__NUM_CONCL__. Conclusiones / Observaciones</div>
        <table class="grid-layout">
            <tr>
                <td>
                    <strong>Observaciones:</strong> __CONCLUSIONES__<br>
                    <strong>Modo Final Validado:</strong> __MODO_FINAL__
                </td>
            </tr>
        </table>
    </div>

    <div class="banner-final-pdf" style="background-color: __COLOR_FINAL__; border: none;">
        <div style="font-size: 8pt; font-weight: bold; color: white; text-transform: uppercase;">Modo Final de INTERVENCIÓN VALIDADO</div>
        <div style="font-size: 14pt; font-weight: 900; color: white;">__MODO_FINAL__</div>
        <div style="font-size: 9pt; font-weight: bold; color: white; margin-top: 5px;">__MENSAJE_SEGURIDAD__</div>
    </div>

    <div style="margin-top: 40px; page-break-inside: avoid;">
        <table style="width: 100%; border-collapse: collapse;">
            <tr>
                <td style="width: 50%; border: 1px solid #cbd5e1; padding: 15px; vertical-align: top;">
                    <strong>Evaluó:</strong><br><br>
                    __EVALUADOR_FULL__
                </td>
                <td style="width: 50%; border: 1px solid #cbd5e1; padding: 15px; vertical-align: top;">
                    <strong>Aprobó:</strong><br><br>
                    <div style="border-bottom: 1px solid black; width: 80%; margin-top: 20px;"></div>
                    <div style="font-size: 8pt; margin-top: 5px;">Firma y Aclaración</div>
                </td>
            </tr>
        </table>
    </div>
</body>
</html>
"""

async def generate_pdf(html_content, output_path):
    async with async_playwright() as p:
        browser = await p.chromium.launch(
    headless=True,
    args=[
        "--no-sandbox",
        "--disable-dev-shm-usage"
    ]
)
        page = await browser.new_page()
        await page.set_content(html_content)
        await page.pdf(path=output_path, format="A4", print_background=True)
        await browser.close()

# --- COMPILACIÓN ---
st.markdown("---")
if st.button("✅ COMPILAR REPORTE (PDF y Excel)"):
    if not lista_peligros_final:
        st.error("Por favor, selecciona al menos un peligro para compilar el informe.")
    else:
        # Mapeo de variables para el reporte
        area_proceso = sitio
        linea_seleccionada = linea
        lista_tareas_final = ", ".join(tareas) if isinstance(tareas, list) else tareas
        
        with st.spinner("Procesando reporte y matriz..."):
            logo_b64 = file_to_base64("arca.png")
            logo_tag = f'<img src="data:image/png;base64,{logo_b64}" style="max-height:45px;">' if logo_b64 else '<h2>ARCA</h2>'
            
            equip_b64 = get_base64_image(equip_file)
            equip_tag = f'<img src="data:image/png;base64,{equip_b64}">' if equip_b64 else '<p style="font-size:7pt; color:#64748b;">[Foto no cargada en sistema]</p>'
            
            arbol_b64 = file_to_base64("arbol.png")
            arbol_tag = f'<img src="data:image/png;base64,{arbol_b64}">' if arbol_b64 else '<p style="font-size:7pt; color:#ef4444;">[Falta archivo fijo arbol.png]</p>'
            
            # Construcción estricta de las filas con todo el desglose analítico de Fine
            html_rows = ""
            for item in evaluaciones_fine:
                _, col_r = clasificar_fine(item["gp_r"])
                _, col_i = clasificar_fine(item["gp_i"])
                
                html_rows += f"""
                <tr>
                    <td class="left-align" style="font-weight:bold; width:18%;">{item['peligro']}</td>
                    <td>{item['p_i']}</td>
                    <td>{item['e_i']}</td>
                    <td>{item['c_i']}</td>
                    <td style="background:#f8fafc; font-weight:bold; color:{col_i};">{item['gp_i']}</td>
                    <td class="left-align" style="color:#B51E2D; font-size:7.5pt; width:25%;">{item['controles'] if item['controles'] else 'Sin mitigantes'}</td>
                    <td>{item['p_r']}</td>
                    <td>{item['e_r']}</td>
                    <td>{item['c_r']}</td>
                    <td style="background:#f8fafc; font-weight:bold; color:{col_r};">{item['gp_r']}</td>
                </tr>
                """
            
            html_f = HTML_TEMPLATE
            html_f = html_f.replace("__LOGO_HTML__", logo_tag).replace("__FOTO_EQUIPO__", equip_tag).replace("__ARBOL_LOTO__", arbol_tag)
            html_f = html_f.replace("__PLANTA__", sitio).replace("__FECHA_ACTUAL__", datetime.datetime.now().strftime("%d/%m/%Y"))
            html_f = html_f.replace("__NEGOCIO__", negocio).replace("__TIPO_SITIO__", tipo_sitio)
            html_f = html_f.replace("__AREA__", area_sector).replace("__LINEA__", linea)
            
            equipo_desc = maquina if area_sector == "Producción" else subsector_equipo
            html_f = html_f.replace("__EQUIPO_DESC__", equipo_desc)
            
            html_f = html_f.replace("__FABRICANTE__", fabricante).replace("__MODELO__", modelo).replace("__ANIO__", str(anio))
            html_f = html_f.replace("__CLASIFICACION__", clasificacion)
            html_f = html_f.replace("__FRECUENCIA__", frecuencia).replace("__DURACION__", duracion)
            
            # Lógica de numeración y validación en PDF
            if modo_inicial in ["Modo 1", "Modo 2"]:
                res_val = ""
                if modo_final == "Modo 3" and modo_inicial != "Modo 3":
                    razon = "el riesgo residual es no aceptable (GP > 20)" if riesgo_no_aceptable else "no se garantiza un control fiable"
                    res_val = f"Como {razon}, entonces el modo convalidado es Modo 3."
                else:
                    res_val = "Como el riesgo residual es aceptable (GP < 20) y se garantiza un control fiable, entonces se mantiene el modo Inicial seleccionado."
                
                val_html = f"""
                <div class="section">
                    <div class="section-header">5. Validación del Modo Inicial</div>
                    <table class="grid-layout">
                        <tr>
                            <td>
                                <strong>{pregunta if 'pregunta' in locals() else '¿Se garantiza control fiable para prevenir rearranque?'}</strong><br>
                                <strong>Respuesta:</strong> {control_fiable if 'control_fiable' in locals() else 'N/A'}<br>
                                <strong>Resultado:</strong> {res_val}
                            </td>
                        </tr>
                    </table>
                </div>
                """
                html_f = html_f.replace("__SECCION_VALIDACION__", val_html)
                html_f = html_f.replace("__NUM_CONCL__", "6")
            else:
                html_f = html_f.replace("__SECCION_VALIDACION__", "")
                html_f = html_f.replace("__NUM_CONCL__", "5")
            
            # Sección 2: Identificación de Modo Inicial
            html_f = html_f.replace("__NIVEL_ACCESO__", nivel_acceso)
            html_f = html_f.replace("__MODO_INICIAL__", modo_inicial)
            
            if modo_inicial in ["Modo 1", "Modo 2"]:
                accion_pdf = "*Modo sujeto a validación en 5. VALIDACIÓN DEL MODO INICIAL"
            else:
                accion_pdf = mensajes_seguridad.get(modo_inicial, "N/A")
            html_f = html_f.replace("__ACCION_REQUERIDA__", accion_pdf)
            
            # Formatear tareas para el PDF (separadas por coma)
            tareas_txt = ", ".join(lista_tareas_combinadas) if lista_tareas_combinadas else "No parametrizadas"
            html_f = html_f.replace("__TAREAS__", tareas_txt)
            
            # Formatear energías para el PDF
            energia_txt = ", ".join([f"{k} ({v})" for k, v in energias_seleccionadas.items()]) if energias_seleccionadas else "Ninguna identificada"
            html_f = html_f.replace("__ENERGIA__", energia_txt)
            
            evaluador_full = f"{evaluador} - {puesto_evaluador} - {fecha_evaluacion.strftime('%d/%m/%Y')}"
            html_f = html_f.replace("__EVALUADOR_FULL__", evaluador_full).replace("__FILAS_MATRIZ__", html_rows)
            html_f = html_f.replace("__CONCLUSIONES__", medidas_conclusiones).replace("__MODO_FINAL__", modo_final.upper())
            html_f = html_f.replace("__COLOR_FINAL__", colores_modos.get(modo_final, "#B51E2D"))
            
            # Añadir mensaje de seguridad dinámico al PDF
            msg_pdf = mensajes_seguridad.get(modo_final, "")
            html_f = html_f.replace("__MENSAJE_SEGURIDAD__", msg_pdf)
            
            pdf_out = f"Ficha_LOTO_Integral_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            asyncio.run(generate_pdf(html_f, pdf_out))
            
            # --- GENERACIÓN DE EXCEL TÉCNICO AVANZADO ---
            from openpyxl.drawing.image import Image as XLImage
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ficha LOTO Integral"
            
            # Configuración de anchos de columna
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['F'].width = 30
            for col in ['B','C','D','E','G','H','I','J']:
                ws.column_dimensions[col].width = 12
            
            # Estilos
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            red_fill = PatternFill(start_color="B51E2D", end_color="B51E2D", fill_type="solid")
            gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            bold_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 1. ENCABEZADO ISO/SOP
            ws.merge_cells('A1:B4') # Espacio para Logo
            if os.path.exists("arca.png"):
                img_logo = XLImage("arca.png")
                img_logo.width = 120
                img_logo.height = 60
                ws.add_image(img_logo, 'A1')
            
            ws.merge_cells('C1:H2')
            ws['C1'] = "CONTROL DE ENERGÍAS PELIGROSAS"
            ws['C1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['C1'].fill = black_fill
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('C3:H4')
            ws['C3'] = f"{negocio} - {sitio} | MODOS DE INTERVENCIÓN Y EVALUACIÓN DE RIESGOS"
            ws['C3'].font = Font(size=11, bold=True, color="FFFFFF")
            ws['C3'].fill = red_fill
            ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('I1:J1'); ws['I1'] = "CÓDIGO: "
            ws.merge_cells('I2:J2'); ws['I2'] = "REVISIÓN: "
            ws.merge_cells('I3:J4'); ws['I3'] = f"FECHA: {datetime.datetime.now().strftime('%d/%m/%Y')}"
            for r in range(1, 5):
                for c in range(9, 11):
                    ws.cell(row=r, column=c).border = border
                    ws.cell(row=r, column=c).font = Font(size=8, bold=True)
            
            ws.append([""]) # Espacio
            
            # 2. IDENTIFICACIÓN
            row = ws.max_row + 1
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value="1. IDENTIFICACIÓN DE EQUIPO Y TAREAS").font = white_font
            ws.cell(row=row, column=1).fill = red_fill
            
            ws.append(["Negocio", negocio, "Sitio", sitio, "", "Ubicación", f"{area_sector} - {linea}"])
            ws.append(["Equipo", equipo_desc, "Fabricante", fabricante, "Modelo", modelo, "Año", anio])
            ws.append(["Clasificación", clasificacion, "Frecuencia", frecuencia, "Duración", duracion])
            
            # Foto del Equipo en Excel
            if equip_file:
                try:
                    img_data = io.BytesIO(equip_file.getvalue())
                    img_equip = XLImage(img_data)
                    img_equip.width = 200
                    img_equip.height = 150
                    ws.add_image(img_equip, f'F{ws.max_row + 2}')
                except: pass
            
            ws.append(["Tareas", tareas_txt])
            ws.merge_cells(f'B{ws.max_row}:E{ws.max_row}')
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)
            
            # 3. MODO INICIAL
            ws.append([""])
            row = ws.max_row + 1
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value="2. IDENTIFICACIÓN DE MODO INICIAL").font = white_font
            ws.cell(row=row, column=1).fill = red_fill
            
            ws.append(["Nivel de Acceso", nivel_acceso]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            ws.append(["Modo Inicial", modo_inicial])
            ws.append(["Acción Requerida", accion_pdf]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            ws.append(["Evaluador", evaluador_full]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            
            # 4. ENERGÍAS
            ws.append([""])
            row = ws.max_row + 1
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value="3. FUENTE DE ENERGÍA(S)").font = white_font
            ws.cell(row=row, column=1).fill = red_fill
            ws.append(["Energías", energia_txt]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            
            # 5. MATRIZ FINE
            ws.append([""])
            row = ws.max_row + 1
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value="4. DESGLOSE DE MATRIZ WILLIAM FINE").font = white_font
            ws.cell(row=row, column=1).fill = red_fill
            
            headers_fine = ["Peligro Identificado", "P(i)", "E(i)", "C(i)", "GP(i)", "Controles Operacionales", "P(r)", "E(r)", "C(r)", "GP(r)"]
            ws.append(headers_fine)
            for cell in ws[ws.max_row]:
                cell.font = white_font
                cell.fill = black_fill
                cell.border = border
                
            for item in evaluaciones_fine:
                ws.append([item['peligro'], item['p_i'], item['e_i'], item['c_i'], item['gp_i'], item['controles'], item['p_r'], item['e_r'], item['c_r'], item['gp_r']])
                ws.cell(row=ws.max_row, column=6).alignment = Alignment(wrap_text=True)
                for cell in ws[ws.max_row]:
                    cell.border = border
            
            # 6. VALIDACIÓN Y CONCLUSIÓN
            ws.append([""])
            if modo_inicial in ["Modo 1", "Modo 2"]:
                row = ws.max_row + 1
                ws.merge_cells(f'A{row}:J{row}')
                ws.cell(row=row, column=1, value="5. VALIDACIÓN DEL MODO INICIAL").font = white_font
                ws.cell(row=row, column=1).fill = red_fill
                ws.append(["Pregunta", pregunta if 'pregunta' in locals() else "N/A"]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
                ws.append(["Respuesta", control_fiable if 'control_fiable' in locals() else "N/A"])
                ws.append(["Resultado", res_val]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
                ws.append([""])
                row = ws.max_row + 1
                ws.merge_cells(f'A{row}:J{row}')
                ws.cell(row=row, column=1, value="6. CONCLUSIÓN Y MODO FINAL").font = white_font
                ws.cell(row=row, column=1).fill = red_fill
            else:
                row = ws.max_row + 1
                ws.merge_cells(f'A{row}:J{row}')
                ws.cell(row=row, column=1, value="5. CONCLUSIÓN Y MODO FINAL").font = white_font
                ws.cell(row=row, column=1).fill = red_fill
            
            ws.append(["Observaciones", medidas_conclusiones]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            ws.append(["Modo Final", modo_final.upper()])
            ws.append(["Instrucciones", msg_pdf]); ws.merge_cells(f'B{ws.max_row}:J{ws.max_row}')
            
            # Pie de Firmas
            ws.append([""]); ws.append([""])
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value="Evaluó:").font = bold_font
            ws.cell(row=row, column=6, value="Aprobó:").font = bold_font
            ws.append([evaluador_full, "", "", "", "", "__________________________"])
            ws.append(["", "", "", "", "", "Firma y Aclaración"])
            
            excel_out = f"Ficha_LOTO_Integral_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(excel_out)
            
            with open(pdf_out, "rb") as f:
                st.session_state["pdf_v5"] = f.read()
            with open(excel_out, "rb") as f:
                st.session_state["excel_v5"] = f.read()
                
            st.success("🎉 Reporte (PDF + Excel) compilado.")

if "pdf_v5" in st.session_state:
    st.download_button("📥 Descargar Reporte en PDF", st.session_state["pdf_v5"], file_name=f"Ficha_LOTO_Integral.pdf", mime="application/pdf")
if "excel_v5" in st.session_state:
    st.download_button("📥 Descargar Reporte en Excel", st.session_state["excel_v5"], file_name=f"Ficha_LOTO_Integral.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
